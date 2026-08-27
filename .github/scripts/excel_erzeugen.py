# -*- coding: utf-8 -*-
"""Erzeugt die Angebotsdaten.xlsx aus der Datenbank.

Die Datei ist seit dem 27.08.2026 nicht mehr die Quelle, sondern der Notweg: die
vier Werkzeuge lesen die Datenbank und greifen nur dann auf sie zurueck, wenn die
Datenbank nicht antwortet. Damit dieser Notweg nicht veraltet, laeuft dieses
Skript taeglich auf GitHub und schreibt die Datei neu.

Zwei Dinge nebenbei:
  • Der Abruf haelt das kostenlose Datenbankprojekt wach, das sonst nach
    laengerer Ruhe stillgelegt wuerde.
  • Der Speicherzeitpunkt der Mappe wird auf den juengsten Aenderungszeitpunkt
    der Daten gesetzt. Die Werkzeuge lesen daraus ihren "Stand" (standAusMappe),
    und die Datei aendert sich dadurch nur, wenn sich die Daten geaendert haben -
    sonst gaebe es jeden Tag einen leeren Commit.

Zugang: Adresse und oeffentlicher Schluessel stehen in konfiguration.js. Lesen ist
offen, ein Geheimnis wird nicht gebraucht.
"""
import io, json, re, sys, urllib.request, zipfile
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

WURZEL = Path(__file__).resolve().parents[2]
ZIEL   = WURZEL / 'Angebotsdaten.xlsx'

PREIS_KOPF = ['Jahrgang', 'Titel', 'Auflage', 'Kategorie', 'Format',
              'Grundpreis / 4c-Preis (€)', '4c-Zuschlag (€)', 'rabattfähig',
              'AE-fähig', 'Termin (DU/EH)', 'Fachbereich', 'Seiten-Äquivalent']
PREIS_BREITE = [10, 40, 10.29, 30, 34, 25.14, 16.57, 13.14, 10.86, 16, 12.86, 17.43]

TERMIN_KOPF = ['Jahrgang', 'Titel', 'Heft', 'Monat', 'ET', 'AS', 'DU-Schluss',
               'EH-Termin', 'Themenschwerpunkte', 'Kongresse']
TERMIN_BREITE = [10, 43.29, 6.86, 10.29, 10.14, 10.29, 13.14, 12.14, 122, 60.43]

INFO = [
    'Angebotsdaten – Stammdaten für die Angebotswerkzeuge',
    None,
    'ACHTUNG: Diese Datei wird täglich aus der Datenbank neu erzeugt. Änderungen',
    'daran erreichen die Werkzeuge nicht und gehen beim nächsten Lauf verloren.',
    'Gepflegt wird in der Preispflege:',
    '   https://barings1995.github.io/angebotswerkzeuge/Preispflege.html',
    'Wer trotzdem in Excel arbeiten möchte, ändert hier und liest die Datei dort',
    'über »Aus Excel laden« wieder ein – dann steht der Stand in der Datenbank.',
    None,
    'Wozu diese Datei dann noch da ist: Sie ist der Notweg. Antwortet die Datenbank',
    'nicht, lesen die Werkzeuge sie und sagen es in der Statuszeile. Außerdem lässt',
    'sie sich per »Daten laden« von Hand öffnen.',
    None,
    'Blatt »Preise« – eine Zeile je Format:',
    '   Jahrgang · Titel · Auflage · Kategorie · Format · Grundpreis / 4c-Preis (€) · 4c-Zuschlag (€) · rabattfähig · AE-fähig · Termin (DU/EH) · Fachbereich · Seiten-Äquivalent',
    '   • Grundpreis / 4c-Preis = s/w-Grundpreis (bei 4c-inklusive der Komplettpreis, 4c-Zuschlag dann leer).',
    '   • rabattfähig / AE-fähig: Pflichtangabe »ja« oder »nein«. Eine leere Zelle wirkt wie »nein«.',
    '   • Termin (DU/EH): »DU« = Druckunterlagenschluss (der Verlag druckt), »EH« = Anliefertermin',
    '     (fertige Ware in der Druckerei). Bleibt die Zelle leer, erscheint kein Produktionstermin.',
    '   • Fachbereich: bündelt die Titel in der Titelauswahl. Je Titel derselbe Wert.',
    '   • Seiten-Äquivalent: nur eintragen, wo der aus dem Formatnamen abgeleitete Wert nicht passt.',
    None,
    'Blatt »Termine« – eine Zeile je Ausgabe:  Jahrgang · Titel · Heft · Monat · ET · AS · DU-Schluss · EH-Termin · Themenschwerpunkte · Kongresse',
    '   • Termine im Format TT.MM. Fällt einer ins Vor- oder Folgejahr, mit Jahreszahl: 22.12.2026.',
    None,
    'Spaltenreihenfolge NICHT ändern – die Werkzeuge lesen nach Position.',
]


def konfiguration():
    text = (WURZEL / 'konfiguration.js').read_text(encoding='utf-8')
    adresse = re.search(r"SUPABASE_URL\s*=\s*'([^']+)'", text).group(1)
    schluessel = re.search(r"SUPABASE_ANON_KEY\s*=\s*'([^']+)'", text).group(1)
    return adresse, schluessel


def holen(adresse, schluessel, pfad):
    anfrage = urllib.request.Request(
        adresse + '/rest/v1/' + pfad,
        headers={'apikey': schluessel, 'Authorization': 'Bearer ' + schluessel})
    with urllib.request.urlopen(anfrage, timeout=60) as antwort:
        return json.loads(antwort.read().decode('utf-8'))


def zahl(wert):
    if wert in (None, ''):
        return None
    z = float(wert)
    return int(z) if z == int(z) else round(z, 2)


def blatt_anlegen(mappe, name, kopf, breiten):
    blatt = mappe.create_sheet(name)
    blatt.append(kopf)
    for zelle in blatt[1]:
        zelle.font = Font(bold=True)
    for i, b in enumerate(breiten, 1):
        blatt.column_dimensions[get_column_letter(i)].width = b
    blatt.freeze_panes = 'A2'
    return blatt


def erzeugen():
    adressen = konfiguration()
    felder = ('id,jahr,reihenfolge,name,auflage,fachbereich,'
              'preis(reihenfolge,kategorie,format,preis,zuschlag_4c,rabattfaehig,'
              'ae_faehig,termin_art,seiten_aequivalent),'
              'termin(reihenfolge,heft,monat,et,anzeigenschluss,du_schluss,eh_termin,themen,kongresse)')
    titel = holen(*adressen, 'titel?select=' + urllib.parse.quote(felder) +
                  '&order=jahr.asc,reihenfolge.asc&preis.order=reihenfolge.asc'
                  '&termin.order=reihenfolge.asc&limit=500')
    stand = holen(*adressen, 'stand?select=geaendert_am')
    geaendert = None
    if stand and stand[0].get('geaendert_am'):
        geaendert = datetime.fromisoformat(stand[0]['geaendert_am'].replace('Z', '+00:00'))

    mappe = openpyxl.Workbook()
    mappe.remove(mappe.active)

    info = mappe.create_sheet('Info')
    info.column_dimensions['A'].width = 82.71
    for zeile in INFO:
        info.append([zeile])
    info['A1'].font = Font(bold=True)

    preise = blatt_anlegen(mappe, 'Preise', PREIS_KOPF, PREIS_BREITE)
    termine = blatt_anlegen(mappe, 'Termine', TERMIN_KOPF, TERMIN_BREITE)

    n_preis = n_termin = 0
    for t in titel:
        for p in t.get('preis') or []:
            preise.append([
                int(t['jahr']), t['name'], int(t['auflage'] or 0), p['kategorie'], p['format'],
                zahl(p['preis']), zahl(p['zuschlag_4c']),
                'ja' if p['rabattfaehig'] else 'nein',
                'ja' if p['ae_faehig'] else 'nein',
                p.get('termin_art') or None, t['fachbereich'] or None,
                zahl(p.get('seiten_aequivalent')),
            ])
            n_preis += 1
        for a in t.get('termin') or []:
            termine.append([
                int(t['jahr']), t['name'], a['heft'], a['monat'] or None,
                a['et'] or None, a['anzeigenschluss'] or None, a['du_schluss'] or None,
                a['eh_termin'] or None, a['themen'] or None, a['kongresse'] or None,
            ])
            n_termin += 1

    # Auswahllisten wie in der bisherigen Datei, damit Eintragen von Hand gefuehrt bleibt
    ja_nein = DataValidation(type='list', formula1='"ja,nein"', allowBlank=True)
    du_eh = DataValidation(type='list', formula1='"DU,EH"', allowBlank=True)
    preise.add_data_validation(ja_nein)
    preise.add_data_validation(du_eh)
    ja_nein.add('H2:I%d' % (n_preis + 1))
    du_eh.add('J2:J%d' % (n_preis + 1))
    preise.auto_filter.ref = 'A1:L%d' % (n_preis + 1)
    termine.auto_filter.ref = 'A1:J%d' % (n_termin + 1)

    # Der Speicherzeitpunkt ist der juengste Aenderungszeitpunkt der Daten. Die
    # Werkzeuge bilden daraus ihren "Stand"; zugleich bleibt die Datei Byte fuer
    # Byte gleich, solange sich nichts geaendert hat.
    if geaendert:
        mappe.properties.modified = geaendert.astimezone(timezone.utc).replace(tzinfo=None)
        mappe.properties.created = geaendert.astimezone(timezone.utc).replace(tzinfo=None)
    mappe.properties.creator = 'Angebotswerkzeuge'
    mappe.properties.lastModifiedBy = 'Angebotswerkzeuge'

    puffer = io.BytesIO()
    mappe.save(puffer)
    return festschreiben(puffer.getvalue(), geaendert), n_preis, n_termin, geaendert


def festschreiben(rohdaten, geaendert):
    """Macht die Datei Byte fuer Byte vorhersagbar.

    Zwei Quellen von Zufall werden ausgeraeumt:

      1. Die Zeitstempel der Archiveintraege. Sie traegen sonst die Uhrzeit der
         Erzeugung, und die Datei waere jeden Tag verschieden - der taegliche Lauf
         erzeugte einen Commit, obwohl sich an den Daten nichts geaendert hat.

      2. dcterms:modified in docProps/core.xml. openpyxl setzt es beim Speichern
         selbst auf die aktuelle Uhrzeit und verwirft, was vorher eingetragen war.
         Das ist doppelt stoerend: die Werkzeuge bilden aus diesem Feld ihren
         "Stand" (standAusMappe), und im Notweg staende dort der Zeitpunkt der
         Erzeugung statt des Alters der Daten. Nachgemessen am 27.08.2026: fuenf
         Laeufe, vier verschiedene Dateien, Unterschied jeweils genau eine Sekunde
         in diesem Feld.
    """
    if geaendert is not None:
        stempel = geaendert.astimezone(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')
        rohdaten = _core_xml_setzen(rohdaten, stempel)

    quelle = zipfile.ZipFile(io.BytesIO(rohdaten))
    puffer = io.BytesIO()
    ziel = zipfile.ZipFile(puffer, 'w', zipfile.ZIP_DEFLATED)
    for eintrag in sorted(quelle.namelist()):
        info = zipfile.ZipInfo(eintrag, date_time=(2026, 1, 1, 0, 0, 0))
        info.compress_type = zipfile.ZIP_DEFLATED
        info.external_attr = 0o600 << 16
        ziel.writestr(info, quelle.read(eintrag))
    ziel.close()
    return puffer.getvalue()


def _core_xml_setzen(rohdaten, stempel):
    """Traegt Erzeugungs- und Aenderungszeitpunkt in docProps/core.xml ein."""
    quelle = zipfile.ZipFile(io.BytesIO(rohdaten))
    puffer = io.BytesIO()
    ziel = zipfile.ZipFile(puffer, 'w', zipfile.ZIP_DEFLATED)
    for eintrag in quelle.namelist():
        inhalt = quelle.read(eintrag)
        if eintrag == 'docProps/core.xml':
            text = inhalt.decode('utf-8')
            for feld in ('dcterms:created', 'dcterms:modified'):
                text, n = re.subn(
                    r'(<%s[^>]*>)[^<]*(</%s>)' % (feld, feld),
                    lambda m: m.group(1) + stempel + m.group(2), text)
                if n != 1:
                    raise RuntimeError('Feld %s in core.xml nicht gefunden' % feld)
            inhalt = text.encode('utf-8')
        ziel.writestr(eintrag, inhalt)
    ziel.close()
    return puffer.getvalue()


def main():
    daten, n_preis, n_termin, geaendert = erzeugen()
    vorher = ZIEL.read_bytes() if ZIEL.exists() else b''
    if daten == vorher:
        print('unveraendert: %d Preise, %d Termine' % (n_preis, n_termin))
        return 0
    ZIEL.write_bytes(daten)
    print('geschrieben: %d Preise, %d Termine, Stand %s'
          % (n_preis, n_termin, geaendert.strftime('%d.%m.%Y %H:%M') if geaendert else '—'))
    return 0


if __name__ == '__main__':
    sys.exit(main())
